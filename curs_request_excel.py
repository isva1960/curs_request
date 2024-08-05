# Скрипт для формирования курса доллара к белорусскому рублю за период.
# Данные формируются в таблицу EXCEL Курс доллара США.xlsx в текущей папке
import openpyxl
import datetime
import requests


def request_curs_val(pdate: str, val: str) -> float:
    ret_val = -1
    try:
        # Делаем запрос. Параметры periodicity=0 - ежедневный курс, ondate=pdate (дата курсов, вводится в программе),
        # parammode=2 - трехзначный буквенный код валюты (ИСО 4217)
        # res - ответ
        res = requests.get(f"https://api.nbrb.by/exrates/rates/{val}",
                           params={"periodicity": 0, "ondate": pdate, "parammode": 2})
    except Exception as e:
        print("Ошибка :", e)
    else:
        date = res.json()
        if 'status' in date:  # Если есть status
            if date['status'] == 404:  # и его значение 404,
                date = ""  # то это означает, что нет данных.
        if date:  # Данные есть
            ret_val = date['Cur_OfficialRate']
            print(ret_val,type(ret_val))
        else:
            print(f"Нет данных по валюте {val}!")
    return ret_val


def main() -> None:
    wb = openpyxl.Workbook()
    wb.iso_dates = True
    ws = wb.active
    ws.title = "Курс доллара США"
    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 7
    ws["A1"] = "Дата курса"
    ws["B1"] = "Курс"
    date_curs = input(
        f"Введите начальную дату в виде ДД.ММ.ГГГГ. ")
    try:
        beg_date = datetime.datetime.strptime(date_curs, '%d.%m.%Y')
    except:
        # Если дата задана неверно, выводим сообщение об ошибке
        print("Начальная дата задана неверно!")
    else:
        date_curs = input(
            f"Введите конечную дату в виде ДД.ММ.ГГГГ. ")
        try:
            end_date = datetime.datetime.strptime(date_curs, '%d.%m.%Y')
        except:
            # Если дата задана неверно, выводим сообщение об ошибке
            print("Конечная дата задана неверно!")
        else:
            if beg_date > end_date:
                print("Конечная дата меньше начальной!")
            else:
                t_date = beg_date
                i = 1
                while t_date <= end_date:
                    i += 1
                    formatted_date = t_date.strftime('%Y-%m-%d')
                    curs_value = request_curs_val(formatted_date, "USD")
                    cell = ws['A' + str(i)]
                    cell.value = t_date
                    cell.number_format = 'DD.MM.YYYY;@'
                    print('Формируются данные за ' + t_date.strftime('%d.%m.%Y'))
                    cell = ws['B' + str(i)]
                    cell.value = curs_value
                    cell.number_format = '0.0000'
                    t_date = t_date + datetime.timedelta(days=1)
                print('Данные сформированы в таблице "Курс доллара США.xlsx"')
                wb.save('Курс доллара США.xlsx')
    input('Нажмите клавишу Enter')


if __name__ == '__main__':
    main()
